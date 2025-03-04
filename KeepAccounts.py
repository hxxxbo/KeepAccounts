import os
import shutil

import pandas as pd
import openpyxl
import readService as rs
import common
from datetime import datetime

if __name__ == '__main__':

    # 加载配置文件和术语
    config, terms = common.load_config()
    f = config['filePath']

    # 获取根目录路径和已处理文件列表路径
    root_dir = f['root_dir']  # 替换为你的目录路径
    processed_file_list_path = f['processed_file']  # 存储已处理文件的文本文件路径

    # 加载已处理文件集合
    processed_files = common.load_processed_files(processed_file_list_path)

    # 获取模板文件路径
    path_template = f['path_template']

    # 初始化用于合并数据的DataFrame
    data_merge = pd.DataFrame()
    rm_merge = pd.DataFrame()

    # 初始化已处理文件计数器
    processed_count = 0

    # 遍历目录及其子目录，获取所有数据
    for dir_path, dirs, files in os.walk(root_dir):
        for file_name in files:
            file_path = os.path.join(dir_path, file_name)

            # 如果文件尚未处理
            if file_path not in processed_files:
                print(f"Processing file: {file_path}")
                # 处理文件
                if "alipay" in file_path:
                    processed_count += 1
                    data, rm = rs.read_zfb(file_path)  # 读取支付宝数据
                elif "微信" in file_path:
                    processed_count += 1
                    data, rm = rs.read_wx(file_path)  # 读取微信数据
                else:
                    continue

                # 合并数据
                data_merge = pd.concat([data_merge, data], axis=0, ignore_index=True)  # 上下拼接合并表格
                rm_merge = pd.concat([rm_merge, rm], axis=0)  # 上下拼接合并表格

                # 标记文件为已处理
                processed_files.add(file_path)

    # 输出需要处理的文件数量
    print(f"Found {processed_count} files to process.\n")

    # 检查是否处理了新文件
    if processed_count == 0:
        print("No new files to process. Exiting the program.\n")
        exit(0)

    # 数据预处理和分类
    data_merge = common.preprocessing(data_merge)
    data_merge = common.classification(data_merge)

    # 格式转换，DataFrame -> List
    merge_list = data_merge.values.tolist()

    # 保存文件
    current_date = datetime.now().strftime('%Y%m%d_%H%M')
    now_file_path = f'{f['out_dir']}/{current_date}_账单.xlsx'

    # 复制模板文件到新文件路径
    shutil.copyfile(path_template, now_file_path)

    # 加载新创建的Excel文件
    workbook = openpyxl.load_workbook(now_file_path)
    sheet = workbook['明细']

    # 删除原有数据
    maxRow = sheet.max_row  # 获取最大行
    for row in reversed(range(2, maxRow + 1)):
        sheet.delete_rows(row)
        maxRow = sheet.max_row  # 获取最大行
    print('\n「明细」 sheet 页原有 ' + str(maxRow) + ' 行数据，将在抹除后写入数据')

    # 写入新数据
    for row in merge_list:
        sheet.append(row)  # openpyxl写文件

    maxRow = sheet.max_row  # 获取最大行
    print('\n「明细」 sheet 页新增 ' + str(maxRow) + ' 行数据')

    # 保存已处理文件列表
    common.save_processed_files(processed_file_list_path, processed_files)

    # 保存无效数据到Excel文件
    with pd.ExcelWriter(f'{f['out_dir']}/{current_date}_无用数据.xlsx') as writer:
        rm_merge.to_excel(writer, sheet_name="无效数据", index=False)

    # 保存工作簿
    workbook.save(now_file_path)

    print("\n成功将数据写入到 " + now_file_path + "运行成功！write successfully!")
    exit(0)