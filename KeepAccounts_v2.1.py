# version:     2.1
# update:      2.1  2021/12/29: 修复支付宝理财收支逻辑bug
# StartTime:   2021/1/6 12:30
# Finished:    2021/1/7 20:30
# Author:      MickLife
# B站:         https://space.bilibili.com/38626658

import pandas as pd
import openpyxl
import datetime
import wxService as wx
import zfbService as zfb
import common

if __name__ == '__main__':

    # # 路径设置
    # print('提示：请在弹窗中选择要导入的【微信】账单文件\n')
    # path_wx = tkinter.filedialog.askopenfilename(title='选择要导入的微信账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    # if path_wx == '':  # 判断是否只导入了微信或支付宝账单中的一个
    #     cancel_wx = 1
    # else:
    #     cancel_wx = 0
    #
    # print('提示：请在弹窗中选择要导入的【支付宝】账单文件\n')
    # path_zfb = tkinter.filedialog.askopenfilename(title='选择要导入的支付宝账单：', filetypes=[('所有文件', '.*'), ('csv文件', '.csv')])
    # if path_zfb == '':  # 判断是否只导入了微信或支付宝账单中的一个
    #     cancel_zfb = 1
    # else:
    #     cancel_zfb = 0
    #
    # while cancel_zfb == 1 and cancel_wx == 1:
    #     print('\n您没有选择任何一个账单！')
    #     exit(1)
    #
    # path_account = tkinter.filedialog.askopenfilename(title='选择要导出的目标账本表格：', filetypes=[('所有文件', '.*'), ('Excel表格', '.xlsx')])
    # while path_account == '':  # 判断是否选择了账本
    #     print('\n年轻人，不选账本怎么记账？')
    #     exit(1)
    #
    # path_write = path_account

    # # 判断是否只导入了微信或支付宝账单中的一个
    # if cancel_wx == 1:
    #     data_wx = pd.DataFrame()
    # else:
    #     data_wx = wx.read_data(path_wx)  # 读数据
    # if cancel_zfb == 1:
    #     data_zfb = pd.DataFrame()
    # else:
    #     data_zfb = zfb.read_data(path_zfb)  # 读数据

    path_wx = '/Users/g01d-01-0712/PycharmProjects/KeepAccounts/files/微信支付账单(20240301-20240331).csv'
    # path_zfb = '/Users/g01d-01-0712/PycharmProjects/KeepAccounts/files/alipay_record_20240522_110105.csv'
    path_zfb = '/Users/g01d-01-0712/PycharmProjects/KeepAccounts/files/alipay_record_20240530_113759.csv'
    path_zfb = '/Users/g01d-01-0712/PycharmProjects/KeepAccounts/files/alipay_record_20240530_134836.csv'
    path_account = '/Users/g01d-01-0712/PycharmProjects/KeepAccounts/files/自动记账2.0_源数据.xlsx'
    path_write = path_account

    data_wx, rm_wx = wx.read_data(path_wx)  # 读数据
    data_zfb, rm_zfb = zfb.read_data(path_zfb)  # 读数据

    data_merge = pd.concat([data_wx, data_zfb], axis=0)  # 上下拼接合并表格
    rm_merge = pd.concat([rm_wx, rm_zfb], axis=0)  # 上下拼接合并表格

    with pd.ExcelWriter('无用数据.xlsx') as writer:
        rm_merge.to_excel(writer, sheet_name='被删除的数据', index=False)

    data_merge = common.preprocessing(data_merge)

    data_class = common.classification(data_merge)

    print("已自动计算乘后金额和交易月份，已合并数据")
    merge_list = data_merge.values.tolist()  # 格式转换，DataFrame->List
    workbook = openpyxl.load_workbook(path_account)  # openpyxl读取账本文件
    sheet = workbook['明细']

    maxRow = sheet.max_row  # 获取最大行

    for row in reversed(range(2, maxRow + 1)):
        sheet.delete_rows(row)
        maxRow = sheet.max_row  # 获取最大行
        # print(row, maxRow)

    maxRow = sheet.max_row  # 获取最大行

    print('\n「明细」 sheet 页已有 ' + str(maxRow) + ' 行数据，将在末尾写入数据')
    for row in merge_list:
        sheet.append(row)  # openpyxl写文件

    # 在最后1行写上导入时间，作为分割线
    now = datetime.datetime.now()
    now = '👆导入时间：' + str(now.strftime('%Y-%m-%d %H:%M:%S'))
    break_lines = [now, '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
    sheet.append(break_lines)
    workbook.save(path_write)  # 保存
    print("\n成功将数据写入到 " + path_write)
    print("\n运行成功！write successfully!")
    exit(1)
